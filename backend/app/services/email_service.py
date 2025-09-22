import os
import base64
import io
from typing import List, Dict, Any, Optional
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import logging

logger = logging.getLogger(__name__)

class EmailService:
    """Service for sending emails with test case attachments using Gmail API"""
    
    SCOPES = ['https://www.googleapis.com/auth/gmail.send']
    
    def __init__(self):
        self.ba_email = os.environ.get("BA_EMAIL", "bhanagearshan@gmail.com")
        self.admin_email = os.environ.get("ADMIN_EMAIL", "arshan.bhanage@sjsu.edu")
        self.service = None
        self._authenticate()
    
    def _authenticate(self):
        """Authenticate with Gmail API"""
        try:
            creds = None
            # The file token.json stores the user's access and refresh tokens.
            token_path = os.path.join(os.path.dirname(__file__), '..', '..', 'token.json')
            credentials_path = os.path.join(os.path.dirname(__file__), '..', '..', 'credentials.json')
            
            if os.path.exists(token_path):
                creds = Credentials.from_authorized_user_file(token_path, self.SCOPES)
            
            # If there are no (valid) credentials available, let the user log in.
            if not creds or not creds.valid:
                if creds and creds.expired and creds.refresh_token:
                    creds.refresh(Request())
                else:
                    # Try to get credentials from environment variables first
                    client_id = os.environ.get("GOOGLE_CLIENT_ID")
                    client_secret = os.environ.get("GOOGLE_CLIENT_SECRET")
                    
                    if client_id and client_secret:
                        # Create credentials from environment variables
                        credentials_info = {
                            "web": {
                                "client_id": client_id,
                                "client_secret": client_secret,
                                "auth_uri": "https://accounts.google.com/o/oauth2/auth",
                                "token_uri": "https://oauth2.googleapis.com/token",
                                "redirect_uris": ["http://localhost:8082/"]
                            }
                        }
                        flow = InstalledAppFlow.from_client_config(credentials_info, self.SCOPES)
                        creds = flow.run_local_server(port=8082)
                    elif os.path.exists(credentials_path):
                        flow = InstalledAppFlow.from_client_secrets_file(credentials_path, self.SCOPES)
                        creds = flow.run_local_server(port=8082)
                    else:
                        logger.warning("Gmail API credentials not found. Email functionality will be limited.")
                        return
                
                # Save the credentials for the next run
                with open(token_path, 'w') as token:
                    token.write(creds.to_json())
            
            self.service = build('gmail', 'v1', credentials=creds)
            logger.info("Gmail API authentication successful")
            
        except Exception as e:
            logger.error(f"Gmail API authentication failed: {str(e)}")
            self.service = None
    
    def generate_excel_file(self, tests: List[Dict[str, Any]]) -> bytes:
        """Generate Excel file from test cases"""
        try:
            # Create workbook and worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Test Cases"
            
            # Define headers for structured format
            headers = [
                "Test Case Name",
                "Preconditions", 
                "Steps",
                "Expected Result",
                "Actual Result",
                "Test Type",
                "Priority",
                "Journey",
                "Requirement Reference",
                "Status",
                "Test Case ID",
                "Created Date"
            ]
            
            # Add headers with styling
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = header_fill
            
            # Add test case data
            for row_idx, test in enumerate(tests, 2):
                # Convert steps list to formatted string
                steps_data = test.get("steps", test.get("test_script", ""))
                if isinstance(steps_data, list):
                    steps_formatted = "\n".join([f"{i+1}. {step}" for i, step in enumerate(steps_data)])
                else:
                    steps_formatted = str(steps_data) if steps_data else ""
                
                # Convert preconditions list to formatted string
                preconditions_data = test.get("preconditions", test.get("precondition_objective", ""))
                if isinstance(preconditions_data, list):
                    preconditions_formatted = "\n".join([f"• {precondition}" for precondition in preconditions_data])
                else:
                    preconditions_formatted = str(preconditions_data) if preconditions_data else ""
                
                # Convert expected_result list to formatted string
                expected_data = test.get("expected_result", test.get("expected", ""))
                if isinstance(expected_data, list):
                    expected_formatted = "\n".join([f"• {exp}" for exp in expected_data])
                else:
                    expected_formatted = str(expected_data) if expected_data else ""
                
                # Use structured format if available, fallback to legacy format
                row_data = [
                    test.get("test_case_name", test.get("name", test.get("title", ""))),
                    preconditions_formatted,
                    steps_formatted,
                    expected_formatted,
                    test.get("actual_result", ""),
                    test.get("test_type", "positive"),
                    test.get("priority", "Medium"),
                    test.get("journey", ""),
                    test.get("requirement_reference", ""),
                    test.get("status", "Draft"),
                    test.get("test_case_id", test.get("key", test.get("test_id", ""))),
                    test.get("created_date", datetime.now().strftime("%Y-%m-%d"))
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col, value=value)
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    
                    # Color code by test type
                    if col == 6:  # Test Type column
                        if value == "positive":
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        elif value == "negative":
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        elif value == "edge":
                            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    
                    # Color code by priority
                    elif col == 7:  # Priority column
                        if value == "High":
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        elif value == "Medium":
                            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        elif value == "Low":
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Add summary sheet
            summary_ws = wb.create_sheet("Summary")
            summary_ws.cell(row=1, column=1, value="Test Case Summary")
            summary_ws.cell(row=1, column=1).font = Font(bold=True, size=16)
            
            # Count test types
            positive_count = sum(1 for test in tests if test.get("test_type") == "positive")
            negative_count = sum(1 for test in tests if test.get("test_type") == "negative")
            edge_count = sum(1 for test in tests if test.get("test_type") == "edge")
            
            summary_data = [
                ["Total Test Cases", len(tests)],
                ["Positive Cases", positive_count],
                ["Negative Cases", negative_count],
                ["Edge Cases", edge_count],
                ["High Priority", sum(1 for test in tests if test.get("priority") == "High")],
                ["Medium Priority", sum(1 for test in tests if test.get("priority") == "Medium")],
                ["Low Priority", sum(1 for test in tests if test.get("priority") == "Low")],
                ["Generated Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
            ]
            
            for row_idx, (label, value) in enumerate(summary_data, 3):
                summary_ws.cell(row=row_idx, column=1, value=label)
                summary_ws.cell(row=row_idx, column=2, value=value)
            
            # Save to BytesIO
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()
            
        except Exception as e:
            logger.error(f"Failed to generate Excel file: {str(e)}")
            raise
    
    def send_test_cases_email(self, tests: List[Dict[str, Any]], journey: str) -> Dict[str, Any]:
        """Send test cases to BA via email with Excel attachment"""
        try:
            if not self.service:
                return {
                    "success": False,
                    "message": "Gmail API not authenticated. Please check credentials."
                }
            
            # Generate Excel file
            excel_data = self.generate_excel_file(tests)
            
            # Create email message
            message = MIMEMultipart()
            message['to'] = self.ba_email
            message['from'] = self.admin_email
            message['subject'] = "Please verify the Test Cases generated"
            
            # Email body
            body = f"""Check the test cases generated attached along with this email.

Journey: {journey}
Total Test Cases: {len(tests)}
Generated Date: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}

Thank You.

Best regards,
TraceQA Admin
"""
            
            message.attach(MIMEText(body, 'plain'))
            
            # Attach Excel file
            attachment = MIMEBase('application', 'octet-stream')
            attachment.set_payload(excel_data)
            encoders.encode_base64(attachment)
            attachment.add_header(
                'Content-Disposition',
                f'attachment; filename=test_cases_{journey}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            )
            message.attach(attachment)
            
            # Encode message
            raw_message = base64.urlsafe_b64encode(message.as_bytes()).decode('utf-8')
            
            # Send email
            send_message = self.service.users().messages().send(
                userId='me',
                body={'raw': raw_message}
            ).execute()
            
            logger.info(f"Email sent successfully. Message ID: {send_message['id']}")
            
            return {
                "success": True,
                "message": "Test cases sent to BA successfully",
                "message_id": send_message['id'],
                "recipient": self.ba_email,
                "attachment_filename": f"test_cases_{journey}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            }
            
        except HttpError as error:
            logger.error(f"Gmail API error: {error}")
            return {
                "success": False,
                "message": f"Gmail API error: {str(error)}"
            }
        except Exception as e:
            logger.error(f"Failed to send email: {str(e)}")
            return {
                "success": False,
                "message": f"Failed to send email: {str(e)}"
            }
    
    def send_simple_email(self, tests: List[Dict[str, Any]], journey: str) -> Dict[str, Any]:
        """Fallback method using Gmail SMTP"""
        try:
            import smtplib
            from email.mime.multipart import MIMEMultipart
            from email.mime.text import MIMEText
            from email.mime.base import MIMEBase
            from email import encoders
            
            # Generate Excel file
            excel_data = self.generate_excel_file(tests)
            
            # Create message
            msg = MIMEMultipart()
            msg['From'] = self.admin_email
            msg['To'] = self.ba_email
            msg['Subject'] = f"Test Cases Generated for {journey} Journey"
            
            # Email body
            body = f"""Dear BA Team,

Please find attached the test cases generated for the {journey} journey.

Journey: {journey}
Total Test Cases: {len(tests)}
Generated Date: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}

The test cases have been generated based on the uploaded requirements documents and are ready for review.

Best regards,
TraceQA System
"""
            
            msg.attach(MIMEText(body, 'plain'))
            
            # Attach Excel file
            attachment = MIMEBase('application', 'octet-stream')
            attachment.set_payload(excel_data)
            encoders.encode_base64(attachment)
            attachment.add_header(
                'Content-Disposition',
                f'attachment; filename=test_cases_{journey}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            )
            msg.attach(attachment)
            
            # For demo purposes, we'll log the email details instead of actually sending
            # In production, you would use:
            # server = smtplib.SMTP('smtp.gmail.com', 587)
            # server.starttls()
            # server.login(self.admin_email, "your_app_password")
            # server.send_message(msg)
            # server.quit()
            
            logger.info(f"Email prepared for: {self.ba_email}")
            logger.info(f"From: {self.admin_email}")
            logger.info(f"Subject: Test Cases Generated for {journey} Journey")
            logger.info(f"Journey: {journey}")
            logger.info(f"Test cases count: {len(tests)}")
            logger.info(f"Excel attachment size: {len(excel_data)} bytes")
            
            return {
                "success": True,
                "message": f"Email prepared successfully for {journey} journey with {len(tests)} test cases",
                "recipient": self.ba_email,
                "journey": journey,
                "test_count": len(tests),
                "attachment_size": len(excel_data),
                "note": "Email details logged (SMTP configured but not sending for demo)"
            }
            
        except Exception as e:
            logger.error(f"Failed to prepare email: {str(e)}")
            return {
                "success": False,
                "message": f"Failed to prepare email: {str(e)}"
            }
